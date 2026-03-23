from fastapi import APIRouter, Depends, HTTPException, Request, Query, UploadFile, File
from fastapi.responses import JSONResponse
from fastapi.templating import Jinja2Templates
from pathlib import Path
from sqlalchemy.orm import Session, joinedload
from typing import Optional
import math
import io
import json

from app.database import get_db
from app.models import Part, BOMHeader, BOMItem, BOMSubstitute
from app.auth import get_current_operator

PAGE_SIZE = 20
bom_router = APIRouter(prefix="/bom")
BASE_DIR = Path(__file__).parent
templates = Jinja2Templates(directory=BASE_DIR / "templates")


# ── 헬퍼 함수 ────────────────────────────────────────────────────────────────

def _ver_key(h):
    try:
        return float(h.version)
    except (ValueError, TypeError):
        return 0.0


def _get_latest_bom(part_id: int, bom_type: str, db: Session):
    """해당 파트+타입의 최신 BOM 헤더 반환"""
    all_boms = db.query(BOMHeader).filter(
        BOMHeader.part_id == part_id,
        BOMHeader.bom_type == bom_type,
    ).all()
    return max(all_boms, key=_ver_key) if all_boms else None


def build_recursive_tree(part_id: int, bom_type: str, db: Session, level: int = 1, version: str = None, visited=None):
    """재귀적 BOM 트리 빌드 (상세 페이지용)"""
    if visited is None:
        visited = set()
    if part_id in visited:
        return []
    visited.add(part_id)

    q = db.query(BOMHeader).filter(BOMHeader.part_id == part_id, BOMHeader.bom_type == bom_type)
    if version:
        bom_header = q.filter(BOMHeader.version == version).first()
    else:
        all_boms = q.all()
        if not all_boms:
            return []
        bom_header = max(all_boms, key=_ver_key)

    if not bom_header:
        return []

    items = (
        db.query(BOMItem)
        .filter(BOMItem.bom_id == bom_header.id)
        .options(
            joinedload(BOMItem.child_part),
            joinedload(BOMItem.substitutes).joinedload(BOMSubstitute.substitute_part),
        )
        .order_by(BOMItem.seq_no)
        .all()
    )

    result = []
    for item in items:
        children = build_recursive_tree(item.child_part_id, bom_type, db, level + 1, visited=set(visited))
        child_bom = None
        if children:
            child_bom_q = db.query(BOMHeader).filter(
                BOMHeader.part_id == item.child_part_id, BOMHeader.bom_type == bom_type,
            ).all()
            if child_bom_q:
                child_bom = max(child_bom_q, key=_ver_key)

        result.append({
            "item": item, "level": level, "children": children,
            "child_bom_header": child_bom,
        })

    return result


def node_to_dict(node):
    """트리 노드를 JSON 직렬화 가능한 dict로 변환"""
    item = node["item"]
    part = item.child_part
    child_bom = node.get("child_bom_header")
    return {
        "id": item.id,
        "part_number": part.part_number if part else "",
        "description": part.description if part else "",
        "spec": part.spec if part else "",
        "category": part.category or "",
        "quantity": item.quantity,
        "unit": item.unit or "",
        "seq_no": item.seq_no,
        "remark": item.remark or "",
        "level": node["level"],
        "child_part_id": item.child_part_id,
        "child_bom_version": child_bom.version if child_bom else None,
        "children": [node_to_dict(c) for c in node["children"]],
        "substitutes": [
            {
                "id": s.id,
                "part_number": s.substitute_part.part_number if s.substitute_part else "",
                "description": s.substitute_part.description if s.substitute_part else "",
                "priority": s.priority,
                "remark": s.remark or "",
            }
            for s in item.substitutes
        ],
    }


def build_editable_tree(part_id: int, bom_type: str, db: Session, level: int = 1, visited=None):
    """편집 페이지용 재귀 트리 (bom_id, check-out 정보 포함)"""
    if visited is None:
        visited = set()
    if part_id in visited:
        return []
    visited.add(part_id)

    bom_header = _get_latest_bom(part_id, bom_type, db)
    if not bom_header:
        return []

    items = (
        db.query(BOMItem)
        .filter(BOMItem.bom_id == bom_header.id)
        .options(
            joinedload(BOMItem.child_part),
            joinedload(BOMItem.substitutes).joinedload(BOMSubstitute.substitute_part),
        )
        .order_by(BOMItem.seq_no)
        .all()
    )

    result = []
    for item in items:
        part_obj = item.child_part
        children = build_editable_tree(item.child_part_id, bom_type, db, level + 1, set(visited))

        child_bom = _get_latest_bom(item.child_part_id, bom_type, db)
        has_own_bom = child_bom is not None

        result.append({
            "id": item.id,
            "bom_id": bom_header.id,
            "child_part_id": item.child_part_id,
            "part_number": part_obj.part_number if part_obj else "",
            "description": part_obj.description if part_obj else "",
            "spec": part_obj.spec if part_obj else "",
            "category": part_obj.category or "" if part_obj else "",
            "unit": item.unit or "EA",
            "quantity": item.quantity,
            "seq_no": item.seq_no,
            "effective_start": item.effective_start or "",
            "effective_end": item.effective_end or "9999-12-31",
            "remark": item.remark or "",
            "level": level,
            "has_own_bom": has_own_bom,
            "child_bom_id": child_bom.id if child_bom else None,
            "child_bom_version": child_bom.version if child_bom else None,
            "child_bom_checked_out": child_bom.checked_out if child_bom else 0,
            "child_bom_checked_out_by": child_bom.checked_out_by if child_bom else None,
            "children": children,
            "substitutes": [
                {
                    "part_id": s.substitute_part_id,
                    "part_number": s.substitute_part.part_number if s.substitute_part else "",
                    "description": s.substitute_part.description if s.substitute_part else "",
                    "priority": s.priority,
                    "remark": s.remark or "",
                }
                for s in item.substitutes
            ],
        })

    return result


def _save_items_to_bom(items_data, bom_header, db):
    """BOM 헤더에 아이템 저장 (기존 삭제 후 새로 생성)"""
    db.query(BOMItem).filter(BOMItem.bom_id == bom_header.id).delete(synchronize_session=False)
    db.flush()

    for idx, item_data in enumerate(items_data):
        child_part_id = item_data.get("child_part_id")
        if not child_part_id:
            continue

        new_item = BOMItem(
            bom_id=bom_header.id,
            child_part_id=child_part_id,
            quantity=float(item_data.get("quantity", 1)),
            unit=item_data.get("unit", "EA") or "EA",
            seq_no=idx,
            effective_start=item_data.get("effective_start") or None,
            effective_end=item_data.get("effective_end") or "9999-12-31",
            remark=item_data.get("remark", "") or "",
        )
        db.add(new_item)
        db.flush()

        for sub_data in item_data.get("substitutes", []):
            sub_part_id = sub_data.get("part_id")
            if not sub_part_id:
                continue
            db.add(BOMSubstitute(
                bom_item_id=new_item.id,
                substitute_part_id=sub_part_id,
                priority=int(sub_data.get("priority", 1)),
                remark=sub_data.get("remark", "") or "",
            ))


def _save_tree_recursive(items_data, parent_part_id, bom_type, db, root_version=None):
    """트리 구조를 재귀적으로 각 BOM 헤더에 저장"""
    # 부모 파트의 BOM 헤더 찾기
    if root_version:
        bom_header = db.query(BOMHeader).filter(
            BOMHeader.part_id == parent_part_id,
            BOMHeader.bom_type == bom_type,
            BOMHeader.version == root_version,
        ).first()
    else:
        bom_header = _get_latest_bom(parent_part_id, bom_type, db)

    if not bom_header:
        return

    # 이 레벨의 아이템 저장
    _save_items_to_bom(items_data, bom_header, db)

    # 자식 레벨 재귀 저장
    for item_data in items_data:
        children = [c for c in item_data.get("children", []) if c.get("child_part_id")]
        child_part_id = item_data.get("child_part_id")
        if children and child_part_id:
            # 자식 파트의 BOM 헤더 찾기/생성
            child_bom = _get_latest_bom(child_part_id, bom_type, db)
            if not child_bom:
                child_bom = BOMHeader(
                    part_id=child_part_id, bom_type=bom_type,
                    version="1.0", status="작성중",
                )
                db.add(child_bom)
                db.flush()
            _save_tree_recursive(children, child_part_id, bom_type, db)


def _cascade_checkout(part_id, bom_type, operator_name, db, visited=None):
    """BOM 체크아웃을 하위 BOM까지 재귀 전파"""
    if visited is None:
        visited = set()
    if part_id in visited:
        return
    visited.add(part_id)

    bom = _get_latest_bom(part_id, bom_type, db)
    if not bom:
        return

    bom.checked_out = 1
    bom.checked_out_by = operator_name
    bom.status = "체크아웃"

    items = db.query(BOMItem).filter(BOMItem.bom_id == bom.id).all()
    for item in items:
        child_bom = _get_latest_bom(item.child_part_id, bom_type, db)
        if child_bom:
            _cascade_checkout(item.child_part_id, bom_type, operator_name, db, visited)


def _cascade_checkin(part_id, bom_type, db, visited=None):
    """BOM 체크인을 하위 BOM까지 재귀 전파 (각각 iteration 증가)"""
    if visited is None:
        visited = set()
    if part_id in visited:
        return None
    visited.add(part_id)

    bom = _get_latest_bom(part_id, bom_type, db)
    if not bom or not bom.checked_out:
        return bom.version if bom else None

    # 하위 BOM 먼저 체크인
    items = db.query(BOMItem).filter(BOMItem.bom_id == bom.id)\
        .options(joinedload(BOMItem.substitutes)).order_by(BOMItem.seq_no).all()
    for item in items:
        child_bom = _get_latest_bom(item.child_part_id, bom_type, db)
        if child_bom and child_bom.checked_out:
            _cascade_checkin(item.child_part_id, bom_type, db, visited)

    # Iteration 증가
    parts = bom.version.split(".")
    major = parts[0]
    iteration = int(parts[1]) + 1 if len(parts) > 1 else 1
    new_version = f"{major}.{iteration}"

    new_bom = BOMHeader(
        part_id=part_id, bom_type=bom_type, version=new_version,
        effective_date=bom.effective_date, status="작성중", checked_out=0,
    )
    db.add(new_bom)
    db.flush()

    for old_item in items:
        new_item = BOMItem(
            bom_id=new_bom.id, child_part_id=old_item.child_part_id,
            quantity=old_item.quantity, unit=old_item.unit, seq_no=old_item.seq_no,
            effective_start=old_item.effective_start, effective_end=old_item.effective_end,
            remark=old_item.remark,
        )
        db.add(new_item)
        db.flush()
        for sub in old_item.substitutes:
            db.add(BOMSubstitute(
                bom_item_id=new_item.id, substitute_part_id=sub.substitute_part_id,
                priority=sub.priority, remark=sub.remark,
            ))

    bom.checked_out = 0
    bom.checked_out_by = None
    bom.status = "체크인"

    return new_version


# ── 1. GET /bom/list ─────────────────────────────────────────────────────────

@bom_router.get("/list")
def bom_list(
    request: Request,
    keyword: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    frame: bool = Query(False),
    db: Session = Depends(get_db),
):
    q = db.query(Part)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(Part.part_number.ilike(kw) | Part.description.ilike(kw))
    if category:
        q = q.filter(Part.category == category)

    total = q.count()
    total_pages = max(1, math.ceil(total / PAGE_SIZE))
    page = max(1, min(page, total_pages))
    parts = q.order_by(Part.created_at.desc()).offset((page - 1) * PAGE_SIZE).limit(PAGE_SIZE).all()

    part_ids = [p.id for p in parts]
    bom_headers = db.query(BOMHeader).filter(BOMHeader.part_id.in_(part_ids)).all() if part_ids else []

    bom_type_map = {}
    for h in bom_headers:
        bom_type_map.setdefault(h.part_id, [])
        if h.bom_type not in bom_type_map[h.part_id]:
            bom_type_map[h.part_id].append(h.bom_type)

    categories = [r[0] for r in db.query(Part.category).distinct().order_by(Part.category).all() if r[0]]

    return templates.TemplateResponse("bom/bom_list.html", {
        "request": request, "parts": parts, "bom_type_map": bom_type_map,
        "keyword": keyword or "", "category": category or "", "categories": categories,
        "page": page, "total_pages": total_pages, "total": total,
        "is_frame": frame, "page_size": PAGE_SIZE,
    })


# ── 2. GET /bom/detail/{part_id} ─────────────────────────────────────────────

@bom_router.get("/detail/{part_id}")
def bom_detail(
    part_id: int, request: Request,
    bom_type: str = Query("E-BOM"),
    version: Optional[str] = Query(None),
    effective_date: Optional[str] = Query(None),
    frame: bool = Query(False),
    db: Session = Depends(get_db),
):
    part = db.query(Part).filter(Part.id == part_id).first()
    if not part:
        raise HTTPException(status_code=404, detail="부품을 찾을 수 없습니다.")

    all_versions_raw = db.query(BOMHeader).filter(
        BOMHeader.part_id == part_id, BOMHeader.bom_type == bom_type,
    ).all()
    all_versions = sorted(all_versions_raw, key=_ver_key, reverse=True)

    bom_header = None
    if all_versions:
        if effective_date:
            candidates = [h for h in all_versions if h.effective_date and h.effective_date <= effective_date]
            bom_header = max(candidates, key=lambda h: (h.effective_date, h.version)) if candidates else all_versions[0]
        elif version is not None:
            bom_header = next((h for h in all_versions if str(h.version) == str(version)), None) or all_versions[0]
        else:
            bom_header = all_versions[0]

    tree = build_recursive_tree(part_id, bom_type, db, version=bom_header.version if bom_header else None) if bom_header else []

    existing_bom_types = [
        r[0] for r in db.query(BOMHeader.bom_type).filter(BOMHeader.part_id == part_id).distinct().all()
    ]

    return templates.TemplateResponse("bom/bom_detail.html", {
        "request": request, "part": part, "bom_header": bom_header,
        "tree": tree, "all_versions": all_versions, "bom_type": bom_type,
        "selected_version": bom_header.version if bom_header else None,
        "effective_date": effective_date or "", "existing_bom_types": existing_bom_types,
        "is_frame": frame,
    })


# ── 3. GET /bom/api/tree/{bom_id} ─────────────────────────────────────────────

@bom_router.get("/api/tree/{bom_id}")
def api_bom_tree(bom_id: int, db: Session = Depends(get_db)):
    bom_header = db.query(BOMHeader).filter(BOMHeader.id == bom_id).first()
    if not bom_header:
        raise HTTPException(status_code=404, detail="BOM을 찾을 수 없습니다.")
    tree = build_recursive_tree(bom_header.part_id, bom_header.bom_type, db, version=bom_header.version)
    return JSONResponse(content=[node_to_dict(n) for n in tree])


# ── 4. GET /bom/edit/{part_id} ────────────────────────────────────────────────

@bom_router.get("/edit/{part_id}")
def bom_edit(
    part_id: int, request: Request,
    bom_type: str = Query("E-BOM"),
    version: Optional[str] = Query(None),
    frame: bool = Query(False),
    db: Session = Depends(get_db),
):
    part = db.query(Part).filter(Part.id == part_id).first()
    if not part:
        raise HTTPException(status_code=404, detail="부품을 찾을 수 없습니다.")

    q = db.query(BOMHeader).filter(BOMHeader.part_id == part_id, BOMHeader.bom_type == bom_type)
    if version is not None:
        bom_header = q.filter(BOMHeader.version == str(version)).first()
    else:
        all_boms = q.all()
        bom_header = max(all_boms, key=_ver_key) if all_boms else None

    if not bom_header:
        bom_header = BOMHeader(part_id=part_id, bom_type=bom_type, version="1.0", status="작성중")
        db.add(bom_header)
        db.commit()
        db.refresh(bom_header)

    # 전체 재귀 트리 (편집용 — bom_id, check-out 정보 포함)
    tree_items = build_editable_tree(part_id, bom_type, db)
    tree_data_json = json.dumps(tree_items, ensure_ascii=False)

    # 현재 로그인한 운영진
    current_operator = getattr(getattr(request, 'state', None), 'current_operator', None)
    operator_name = current_operator.name if current_operator else "알 수 없음"

    return templates.TemplateResponse("bom/bom_edit.html", {
        "request": request, "part": part, "bom": bom_header,
        "bom_type": bom_type, "tree_data_json": tree_data_json,
        "is_frame": frame, "operator_name": operator_name,
    })


# ── 5. POST /bom/edit/{part_id}/save ─────────────────────────────────────────

@bom_router.post("/edit/{part_id}/save")
async def bom_save(part_id: int, request: Request, db: Session = Depends(get_db)):
    import traceback
    try:
        body = await request.json()
        bom_type = body.get("bom_type", "E-BOM")
        version = str(body.get("version", "1.0"))
        effective_date = body.get("effective_date") or None
        status = body.get("status", "작성중")
        items_data = body.get("items", [])

        part = db.query(Part).filter(Part.id == part_id).first()
        if not part:
            return JSONResponse(status_code=404, content={"error": "부품을 찾을 수 없습니다."})

        bom_header = db.query(BOMHeader).filter(
            BOMHeader.part_id == part_id, BOMHeader.bom_type == bom_type, BOMHeader.version == version,
        ).first()

        if not bom_header:
            bom_header = BOMHeader(
                part_id=part_id, bom_type=bom_type, version=version,
                effective_date=effective_date, status=status,
            )
            db.add(bom_header)
            db.flush()
        else:
            bom_header.effective_date = effective_date
            bom_header.status = status

        # child_part_id가 없는 아이템 필터링 (새로 추가 후 부품번호 미입력 상태)
        def filter_valid_items(items):
            return [item for item in items if item.get("child_part_id")]

        valid_items = filter_valid_items(items_data)

        # 재귀적으로 트리 전체 저장
        _save_tree_recursive(valid_items, part_id, bom_type, db, root_version=version)
        db.commit()
        return JSONResponse(content={"ok": True, "bom_id": bom_header.id, "version": bom_header.version})
    except Exception as e:
        db.rollback()
        print(f"[BOM SAVE ERROR] part_id={part_id}: {e}")
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": f"저장 실패: {str(e)}"})


# ── 6. POST /bom/edit/{part_id}/upload-excel ─────────────────────────────────

@bom_router.post("/edit/{part_id}/upload-excel")
async def bom_upload_excel(part_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    from openpyxl import load_workbook

    if not file.filename or not file.filename.endswith(".xlsx"):
        return JSONResponse(status_code=400, content={"error": ".xlsx 파일만 업로드 가능합니다."})

    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as e:
        return JSONResponse(status_code=400, content={"error": f"파일 파싱 실패: {str(e)}"})

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return JSONResponse(content={"success": 0, "fail": 0, "errors": [], "items": []})

    header_row = [str(c).strip() if c else "" for c in rows[0]]
    col_map = {}
    for i, h in enumerate(header_row):
        hl = h.lower()
        if "부품번호" in hl or "part" in hl or "partno" in hl:
            col_map["part_number"] = i
        elif "수량" in hl or "qty" in hl or "quantity" in hl:
            col_map["quantity"] = i
        elif "단위" in hl or "unit" in hl:
            col_map["unit"] = i
        elif "비고" in hl or "remark" in hl:
            col_map["remark"] = i

    col_map.setdefault("part_number", 0)
    col_map.setdefault("quantity", 1)
    col_map.setdefault("unit", 2)
    col_map.setdefault("remark", 3)

    success_count = 0
    fail_count = 0
    errors = []
    items = []

    for row_idx, row in enumerate(rows[1:], start=2):
        def get_cell(key):
            ci = col_map.get(key)
            if ci is None or ci >= len(row):
                return None
            return row[ci]

        raw_pn = get_cell("part_number")
        if not raw_pn:
            continue
        part_number = str(raw_pn).strip()
        if not part_number:
            continue

        try:
            qty = float(str(get_cell("quantity")).strip()) if get_cell("quantity") else 1.0
        except Exception:
            qty = 1.0

        unit = str(get_cell("unit")).strip() if get_cell("unit") else "EA"
        remark = str(get_cell("remark")).strip() if get_cell("remark") else ""

        part_obj = db.query(Part).filter(Part.part_number == part_number).first()
        if not part_obj:
            fail_count += 1
            errors.append(f"행 {row_idx}: 부품번호 '{part_number}' 없음")
            continue

        items.append({
            "child_part_id": part_obj.id,
            "part_number": part_obj.part_number,
            "description": part_obj.description,
            "spec": part_obj.spec or "",
            "category": part_obj.category or "",
            "unit": unit or part_obj.unit or "EA",
            "quantity": qty,
            "seq_no": len(items),
            "remark": remark,
            "has_own_bom": False,
            "children": [],
            "substitutes": [],
        })
        success_count += 1

    return JSONResponse(content={"success": success_count, "fail": fail_count, "errors": errors, "items": items})


# ── 7. GET /bom/api/part-info/{part_number} ───────────────────────────────────

@bom_router.get("/api/part-info/{part_number:path}")
def api_part_info(part_number: str, db: Session = Depends(get_db)):
    part = db.query(Part).filter(Part.part_number == part_number).first()
    if not part:
        return JSONResponse(content={"found": False})
    return JSONResponse(content={
        "found": True, "id": part.id, "part_number": part.part_number,
        "description": part.description, "spec": part.spec or "",
        "category": part.category or "", "unit": part.unit or "EA",
    })


# ── 8. POST /bom/edit/{part_id}/new-version ──────────────────────────────────

@bom_router.post("/edit/{part_id}/new-version")
async def bom_new_version(part_id: int, request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    bom_type = body.get("bom_type", "E-BOM")
    current_version = str(body.get("version", "1.0"))

    source_bom = db.query(BOMHeader).filter(
        BOMHeader.part_id == part_id, BOMHeader.bom_type == bom_type, BOMHeader.version == current_version,
    ).first()
    if not source_bom:
        return JSONResponse(status_code=404, content={"error": "원본 BOM을 찾을 수 없습니다."})

    parts_v = source_bom.version.split(".")
    major = parts_v[0]
    iteration = int(parts_v[1]) + 1 if len(parts_v) > 1 else 1
    new_version_no = f"{major}.{iteration}"

    new_bom = BOMHeader(
        part_id=part_id, bom_type=bom_type, version=new_version_no,
        effective_date=source_bom.effective_date, status="작성중",
    )
    db.add(new_bom)
    db.flush()

    source_items = db.query(BOMItem).filter(BOMItem.bom_id == source_bom.id)\
        .options(joinedload(BOMItem.substitutes)).order_by(BOMItem.seq_no).all()

    for old_item in source_items:
        new_item = BOMItem(
            bom_id=new_bom.id, child_part_id=old_item.child_part_id,
            quantity=old_item.quantity, unit=old_item.unit, seq_no=old_item.seq_no,
            effective_start=old_item.effective_start, effective_end=old_item.effective_end,
            remark=old_item.remark,
        )
        db.add(new_item)
        db.flush()
        for sub in old_item.substitutes:
            db.add(BOMSubstitute(
                bom_item_id=new_item.id, substitute_part_id=sub.substitute_part_id,
                priority=sub.priority, remark=sub.remark,
            ))

    db.commit()
    return JSONResponse(content={"ok": True, "version": new_version_no})


# ── 9. POST /bom/edit/{part_id}/checkout ─────────────────────────────────────

@bom_router.post("/edit/{part_id}/checkout")
async def bom_checkout(part_id: int, request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    bom_type = body.get("bom_type", "E-BOM")
    version = str(body.get("version", "1.0"))

    bom = db.query(BOMHeader).filter(
        BOMHeader.part_id == part_id, BOMHeader.bom_type == bom_type, BOMHeader.version == version,
    ).first()
    if not bom:
        return JSONResponse(status_code=404, content={"error": "BOM을 찾을 수 없습니다."})
    if bom.checked_out:
        return JSONResponse(status_code=400, content={"error": "이미 체크아웃 상태입니다."})

    # 현재 운영진 이름 가져오기
    current_operator = getattr(getattr(request, 'state', None), 'current_operator', None)
    operator_name = current_operator.name if current_operator else "알 수 없음"

    # 재귀 체크아웃 (하위 BOM 포함)
    _cascade_checkout(part_id, bom_type, operator_name, db)
    db.commit()
    return JSONResponse(content={"ok": True, "version": bom.version, "checked_out_by": operator_name})


# ── 10. POST /bom/edit/{part_id}/checkin ──────────────────────────────────────

@bom_router.post("/edit/{part_id}/checkin")
async def bom_checkin(part_id: int, request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    bom_type = body.get("bom_type", "E-BOM")
    version = str(body.get("version", "1.0"))

    bom = db.query(BOMHeader).filter(
        BOMHeader.part_id == part_id, BOMHeader.bom_type == bom_type, BOMHeader.version == version,
    ).first()
    if not bom:
        return JSONResponse(status_code=404, content={"error": "BOM을 찾을 수 없습니다."})
    if not bom.checked_out:
        return JSONResponse(status_code=400, content={"error": "체크아웃 상태가 아닙니다."})

    # 재귀 체크인 (하위 BOM 포함, 각각 iteration 증가)
    new_version = _cascade_checkin(part_id, bom_type, db)
    db.commit()
    return JSONResponse(content={"ok": True, "version": new_version})


# ── 11. POST /bom/edit/{part_id}/revise ──────────────────────────────────────

@bom_router.post("/edit/{part_id}/revise")
async def bom_revise(part_id: int, request: Request, db: Session = Depends(get_db)):
    """Revise — Major 버전 증가 (1.x → 2.0)"""
    body = await request.json()
    bom_type = body.get("bom_type", "E-BOM")
    version = str(body.get("version", "1.0"))

    bom = db.query(BOMHeader).filter(
        BOMHeader.part_id == part_id, BOMHeader.bom_type == bom_type, BOMHeader.version == version,
    ).first()
    if not bom:
        return JSONResponse(status_code=404, content={"error": "BOM을 찾을 수 없습니다."})
    if bom.status != "승인":
        return JSONResponse(status_code=400, content={"error": "승인 상태에서만 Revise 가능합니다."})

    major = int(bom.version.split(".")[0]) + 1
    new_version = f"{major}.0"

    new_bom = BOMHeader(
        part_id=part_id, bom_type=bom_type, version=new_version,
        effective_date=bom.effective_date, status="작성중", checked_out=0,
    )
    db.add(new_bom)
    db.flush()

    source_items = db.query(BOMItem).filter(BOMItem.bom_id == bom.id)\
        .options(joinedload(BOMItem.substitutes)).order_by(BOMItem.seq_no).all()

    for old_item in source_items:
        new_item = BOMItem(
            bom_id=new_bom.id, child_part_id=old_item.child_part_id,
            quantity=old_item.quantity, unit=old_item.unit, seq_no=old_item.seq_no,
            effective_start=old_item.effective_start, effective_end=old_item.effective_end,
            remark=old_item.remark,
        )
        db.add(new_item)
        db.flush()
        for sub in old_item.substitutes:
            db.add(BOMSubstitute(
                bom_item_id=new_item.id, substitute_part_id=sub.substitute_part_id,
                priority=sub.priority, remark=sub.remark,
            ))

    db.commit()
    return JSONResponse(content={"ok": True, "version": new_version})
