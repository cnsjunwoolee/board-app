from fastapi import APIRouter, Depends, Form, HTTPException, Request, UploadFile, File
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pathlib import Path
from sqlalchemy.orm import Session
from typing import Optional
import uuid
import math
import io

from app.database import get_db
from app.models import Member

PAGE_SIZE = 20

member_router = APIRouter(prefix="/members")

BASE_DIR = Path(__file__).parent
templates = Jinja2Templates(directory=BASE_DIR / "templates")

UPLOAD_DIR = BASE_DIR / "static/uploads"
UPLOAD_DIR.mkdir(exist_ok=True)


# 1. GET /members/ - 회원 목록 (검색 + 페이징)
@member_router.get("/")
def list_members(
    request: Request,
    name: Optional[str] = None,
    grade: Optional[str] = None,
    page: int = 1,
    frame: bool = False,
    db: Session = Depends(get_db),
):
    query = db.query(Member)

    if name:
        query = query.filter(Member.name.contains(name))
    if grade:
        query = query.filter(Member.grade == grade)

    total = query.count()
    total_pages = max(1, math.ceil(total / PAGE_SIZE))
    page = max(1, min(page, total_pages))

    members = (
        query.order_by(Member.created_at.desc())
        .offset((page - 1) * PAGE_SIZE)
        .limit(PAGE_SIZE)
        .all()
    )

    return templates.TemplateResponse(
        "members/list.html",
        {
            "request": request,
            "members": members,
            "total": total,
            "page": page,
            "total_pages": total_pages,
            "search_name": name or "",
            "search_grade": grade or "",
            "is_frame": frame,
        },
    )


# 2. GET /members/new - 회원 등록 폼
@member_router.get("/new")
def new_member_form(request: Request, frame: bool = False):
    return templates.TemplateResponse(
        "members/form.html",
        {"request": request, "member": None, "is_frame": frame},
    )


# 3. POST /members/ - 회원 등록 처리 (사진 업로드 포함)
@member_router.post("/")
async def create_member(
    request: Request,
    name: str = Form(...),
    birth_date: Optional[str] = Form(None),
    grade: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    email: Optional[str] = Form(None),
    score: Optional[int] = Form(None),
    memo: Optional[str] = Form(None),
    photo: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
):
    photo_path = None

    if photo and photo.filename:
        ext = Path(photo.filename).suffix
        unique_name = f"{uuid.uuid4().hex}_{photo.filename}"
        save_path = UPLOAD_DIR / unique_name
        content = await photo.read()
        with open(save_path, "wb") as f:
            f.write(content)
        photo_path = f"uploads/{unique_name}"

    member = Member(
        name=name,
        birth_date=birth_date or None,
        grade=grade or None,
        phone=phone or None,
        email=email or None,
        photo_path=photo_path,
        score=score if score is not None else 0,
        memo=memo or None,
    )
    db.add(member)
    db.commit()
    db.refresh(member)
    return RedirectResponse(url=f"/members/{member.id}", status_code=303)


# 8. GET /members/export - 회원 엑셀 다운로드
@member_router.get("/export")
def export_members(
    name: Optional[str] = None,
    grade: Optional[str] = None,
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    query = db.query(Member)
    if name:
        query = query.filter(Member.name.contains(name))
    if grade:
        query = query.filter(Member.grade == grade)
    members = query.order_by(Member.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "회원목록"

    # 헤더
    headers = ["번호", "이름", "생년월일", "학년", "연락처", "이메일", "평가점수", "메모", "등록일"]
    header_fill = PatternFill(start_color="A50034", end_color="A50034", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 데이터
    for i, m in enumerate(members, 2):
        ws.cell(row=i, column=1, value=m.id)
        ws.cell(row=i, column=2, value=m.name)
        ws.cell(row=i, column=3, value=m.birth_date or "")
        ws.cell(row=i, column=4, value=m.grade or "")
        ws.cell(row=i, column=5, value=m.phone or "")
        ws.cell(row=i, column=6, value=m.email or "")
        ws.cell(row=i, column=7, value=f"{'★' * m.score}{'☆' * (5 - m.score)}")
        ws.cell(row=i, column=8, value=m.memo or "")
        ws.cell(row=i, column=9, value=m.created_at.strftime("%Y-%m-%d"))

    # 열 너비 조정
    widths = [8, 12, 14, 10, 16, 25, 14, 30, 14]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"members_{grade or 'all'}.xlsx" if grade else "members_all.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# 4. GET /members/{id} - 회원 상세
@member_router.get("/{member_id}")
def get_member(member_id: int, request: Request, frame: bool = False, db: Session = Depends(get_db)):
    member = db.query(Member).filter(Member.id == member_id).first()
    if member is None:
        raise HTTPException(status_code=404, detail="회원을 찾을 수 없습니다.")
    return templates.TemplateResponse(
        "members/detail.html",
        {"request": request, "member": member, "is_frame": frame},
    )


# 5. GET /members/{id}/edit - 회원 수정 폼
@member_router.get("/{member_id}/edit")
def edit_member_form(member_id: int, request: Request, frame: bool = False, db: Session = Depends(get_db)):
    member = db.query(Member).filter(Member.id == member_id).first()
    if member is None:
        raise HTTPException(status_code=404, detail="회원을 찾을 수 없습니다.")
    return templates.TemplateResponse(
        "members/form.html",
        {"request": request, "member": member, "is_frame": frame},
    )


# 6. POST /members/{id}/edit - 회원 수정 처리
@member_router.post("/{member_id}/edit")
async def update_member(
    member_id: int,
    request: Request,
    name: str = Form(...),
    birth_date: Optional[str] = Form(None),
    grade: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    email: Optional[str] = Form(None),
    score: Optional[int] = Form(None),
    memo: Optional[str] = Form(None),
    photo: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
):
    member = db.query(Member).filter(Member.id == member_id).first()
    if member is None:
        raise HTTPException(status_code=404, detail="회원을 찾을 수 없습니다.")

    if photo and photo.filename:
        unique_name = f"{uuid.uuid4().hex}_{photo.filename}"
        save_path = UPLOAD_DIR / unique_name
        content = await photo.read()
        with open(save_path, "wb") as f:
            f.write(content)
        member.photo_path = f"uploads/{unique_name}"

    member.name = name
    member.birth_date = birth_date or None
    member.grade = grade or None
    member.phone = phone or None
    member.email = email or None
    member.score = score if score is not None else 0
    member.memo = memo or None

    db.commit()
    db.refresh(member)
    return RedirectResponse(url=f"/members/{member.id}", status_code=303)


# 7. POST /members/{id}/delete - 회원 삭제
@member_router.post("/{member_id}/delete")
def delete_member(member_id: int, db: Session = Depends(get_db)):
    member = db.query(Member).filter(Member.id == member_id).first()
    if member is None:
        raise HTTPException(status_code=404, detail="회원을 찾을 수 없습니다.")
    db.delete(member)
    db.commit()
    return RedirectResponse(url="/members/", status_code=303)
